import pytesseract
import pyautogui
import cv2
import numpy as np
import time
import pandas as pd
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import messagebox, scrolledtext
from openpyxl import load_workbook
import shutil
import tempfile
import os
import pygetwindow as gw
import csv
import sys
from openpyxl.utils.dataframe import dataframe_to_rows

def definir_competencia():
    def on_ok():
        nonlocal competencia, codigo_empresa
        competencia = entry_comp.get()
        codigo_empresa = entry_cod.get()
        dialog.destroy()
        root.quit()

    def on_timeout():
        nonlocal timeout
        timeout = True
        dialog.destroy()
        root.quit()

    # Lê empresas do CSV
    empresas = []
    try:
        with open(r"C:\projeto\empresas.csv", newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                empresas.append({'codigo': row['codigo'], 'nome': row['nome']})
    except Exception as e:
        empresas = []

    competencia = None
    codigo_empresa = None
    timeout = False

    root = tk.Tk()
    root.withdraw()

    dialog = tk.Toplevel(root)
    dialog.title("Definir Competência e Empresa")
    dialog.geometry("500x430")
    dialog.attributes('-topmost', True)

    label1 = tk.Label(dialog, text="Digite a competência no formato mmaaaa (ex: 052025):")
    label1.pack(pady=5)
    entry_comp = tk.Entry(dialog)
    entry_comp.pack(pady=5)
    entry_comp.focus()

    label2 = tk.Label(dialog, text="Digite o código da empresa:")
    label2.pack(pady=5)
    entry_cod = tk.Entry(dialog)
    entry_cod.pack(pady=5)

    label_search = tk.Label(dialog, text="Buscar empresa pelo nome:")
    label_search.pack(pady=5)
    entry_search = tk.Entry(dialog)
    entry_search.pack(pady=5)

    label3 = tk.Label(dialog, text="Códigos e empresas disponíveis:")
    label3.pack(pady=5)
    listbox = tk.Listbox(dialog, height=10, width=60)
    listbox.pack(pady=5, fill=tk.BOTH, expand=True)

    def update_listbox(filter_text=""):
        listbox.delete(0, tk.END)
        filtered = [
            emp for emp in empresas
            if filter_text.lower() in emp['nome'].lower()
        ] if filter_text else empresas
        for emp in filtered:
            listbox.insert(tk.END, f"{emp['codigo']} - {emp['nome']}")

    def on_listbox_select(event):
        selection = listbox.curselection()
        if selection:
            valor = listbox.get(selection[0])
            codigo = valor.split(" - ")[0]
            entry_cod.delete(0, tk.END)
            entry_cod.insert(0, codigo)

    def on_search_change(event):
        filter_text = entry_search.get()
        update_listbox(filter_text)

    listbox.bind('<<ListboxSelect>>', on_listbox_select)
    entry_search.bind('<KeyRelease>', on_search_change)

    update_listbox()

    # Botão OK (Confirmar escolha)
    btn = tk.Button(dialog, text="OK (Confirmar escolha)", command=on_ok, bg="#4CAF50", fg="white", font=("Arial", 12, "bold"))
    btn.pack(pady=15)

    dialog.after(20000, on_timeout)
    root.mainloop()

    if timeout:
        return None, None

    # Validação simples
    if (
        competencia and len(competencia) == 6 and competencia[:2].isdigit() and competencia[2:].isdigit()
        and codigo_empresa and codigo_empresa.isdigit()
    ):
        mes = int(competencia[:2])
        if 1 <= mes <= 12:
            return competencia, codigo_empresa
    return None, None

competencia, codigo_empresa = definir_competencia()
if competencia is None or codigo_empresa is None:
    print("A competência ou o código da empresa não foi definido ou o tempo expirou. Encerrando o programa.")
    sys.exit(0)

print(f"Competência escolhida: {competencia}")
print(f"Código da empresa escolhido: {codigo_empresa}")
time.sleep(5)
sys.stdout.reconfigure(encoding='utf-8')

def abrir_janela(competencia, codigo_empresa, target_title="Ambiente Contábil ÚNICO"):
    windows = gw.getWindowsWithTitle(target_title)

    if windows:
        selected_window = windows[0]
        if selected_window.isMinimized:
            selected_window.restore()
        selected_window.activate()
        print(f"Janela '{selected_window.title}' ativada!")
        time.sleep(2)
        pyautogui.hotkey('ctrl', '2')
        time.sleep(8)
        pyautogui.hotkey('ctrl','b')
        time.sleep(1)
        pyautogui.press('esc')
        time.sleep(1)
        pyautogui.write(codigo_empresa)
        time.sleep(3)
        pyautogui.press('enter')
        time.sleep(1)
        pyautogui.write('01'+competencia)
        time.sleep(2)
        pos = pyautogui.locateOnScreen(r"C:\\projeto\\robo\\icones_bal\\export_bal-1.png", confidence=0.70)
        if pos is None:
            print("Ícone de exportação não encontrado. Verifique se a janela está correta.")
            return False
        pyautogui.moveTo(pos)
        time.sleep(1)
        pyautogui.click(button='left')
        pos2 = pyautogui.locateOnScreen(r"C:\\projeto\\robo\\icones_bal\\export_bal-2.png", confidence=0.7)
        if pos2 is None:
            print("Ícone de exportação não encontrado. Verifique se a janela está correta.")
            return False
        pyautogui.moveTo(pos2)
        time.sleep(1)
        pyautogui.click(button='left')
        time.sleep(2)
        pyautogui.press('enter')
        time.sleep(2)
        pyautogui.press('enter')
        time.sleep(2)
        pyautogui.write(caminho_balancete)
        time.sleep(2)
        pyautogui.press('enter')
                                 
caminho_balancete = f'C:\\projeto\\planilhas\\balancete\\balancete_{competencia}_{codigo_empresa}'   
abrir_janela(competencia, codigo_empresa)

def processar_balancete_csv(competencia, codigo_empresa, timeout=30):

    competencia = competencia.strip()
    codigo_empresa = codigo_empresa.strip()
    caminho_balancete = os.path.join(
        'C:\\projeto\\planilhas\\balancete',
        f'balancete_{competencia}_{codigo_empresa}.csv'
    )
    print(f"Tentando abrir: {caminho_balancete}")

    # Busca o nome da empresa no empresas.csv
    nome_empresa = None
    try:
        with open(r"C:\projeto\empresas.csv", newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                if row['codigo'].strip() == codigo_empresa:
                    nome_empresa = row['nome'].strip()
                    break
    except Exception as e:
        print(f"Erro ao ler empresas.csv: {e}")
        return

    if not nome_empresa:
        print(f"Código da empresa {codigo_empresa} não encontrado em empresas.csv.")
        return

    # Gera nome de arquivo seguro
    nome_empresa_arquivo = nome_empresa.replace(' ', '_').replace('/', '_').replace('\\', '_')

    novo_caminho_balancete = os.path.join(
        'C:\\projeto\\planilhas\\balancete',
        f'balancete_{competencia}_{nome_empresa_arquivo}.csv'
    )

    # Aguarda até o arquivo existir ou até o timeout
    tempo_inicial = time.time()
    while not os.path.exists(caminho_balancete):
        if time.time() - tempo_inicial > timeout:
            print(f"Arquivo não encontrado após {timeout} segundos.")
            return
        time.sleep(1)

    try:
        df = pd.read_csv(caminho_balancete, encoding='latin1', sep=';')
        if '259' in df.iloc[:, 0].astype(str).values:
            linha_259 = df[df.iloc[:, 0].astype(str) == '259'].index[0]
            df = df.iloc[:linha_259 + 1, :]
            df.to_csv(novo_caminho_balancete, index=False, sep=';', encoding='latin1')
            print(f"Arquivo processado e salvo como: {novo_caminho_balancete}")
            print(df.head())
            # Remove o arquivo original
            try:
                os.remove(caminho_balancete)
            except Exception as e:
                print(f"Não foi possível remover o arquivo original: {e}")
        else:
            print("String '259' não encontrada na primeira coluna.")
    except Exception as e:
        print(f"Erro ao processar o balancete: {e}")

processar_balancete_csv(competencia, codigo_empresa)

# --- Obtém o nome da empresa para passar à função ---
def get_nome_empresa(codigo_empresa):
    try:
        with open(r"C:\projeto\empresas.csv", newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                if row['codigo'].strip() == codigo_empresa:
                    return row['nome'].strip()
    except Exception as e:
        print(f"Erro ao ler empresas.csv: {e}")
    return None

def colar_balancete_no_excel(competencia, nome_empresa):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    nome_empresa_arquivo = nome_empresa.replace(' ', '_').replace('/', '_').replace('\\', '_')
    caminho_csv = f'C:\\projeto\\planilhas\\balancete\\balancete_{competencia}_{nome_empresa_arquivo}.csv'
    caminho_modelo = r'C:\projeto\planilhas\CONCILIACAO_EMPRESA_XX_XXXX.xlsx'
    caminho_saida = f'C:\\projeto\\planilhas\\CONCILIACAO_{nome_empresa_arquivo}_{competencia}.xlsx'

    # Lê o CSV e cria o DataFrame df
    try:
        df = pd.read_csv(caminho_csv, encoding='latin1', sep=';')
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    except Exception as e:
        print(f"Erro ao ler o balancete processado: {e}")
        return

    try:
        wb = load_workbook(caminho_modelo)
        ws = wb.active
    except Exception as e:
        print(f"Erro ao abrir o modelo Excel: {e}")
        return

    # Inserir os dados começando na coluna A
    start_column = 1  # Coluna A
    start_row = 1
    
    # Agora df está definido e pode ser usado aqui
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, start_column):
            try:
                cell = ws.cell(row=r_idx, column=c_idx)
                # Verifica se a célula não é None e se pode receber valor
                if cell is not None:
                    # Tenta escrever o valor, lidando com células mescladas
                    try:
                        # Para células mescladas, tenta acessar a célula principal do merge
                        if hasattr(cell, 'value'):
                            cell.value = value
                        else:
                            # Se não conseguir acessar, pula
                            print(f"Pulando célula sem acesso na linha {r_idx}, coluna {c_idx}")
                            continue
                    except (AttributeError, ValueError) as e:
                        # Se der erro, pode ser célula mesclada ou protegida
                        print(f"Pulando célula com restrição na linha {r_idx}, coluna {c_idx}: {e}")
                        continue
            except Exception as e:
                print(f"Erro ao escrever na célula ({r_idx}, {c_idx}): {e}")
                continue

    try:
        wb.save(caminho_saida)
        print(f"Arquivo salvo em: {caminho_saida}")
    except Exception as e:
        print(f"Erro ao salvar o arquivo Excel: {e}")

nome_empresa = get_nome_empresa(codigo_empresa)
if nome_empresa:
    colar_balancete_no_excel(competencia, nome_empresa)
else:
    print("Nome da empresa não encontrado.")

# # # Função para processar o balancete baixado
# # def processar_balancete(caminho_balancete, codigo_empresa, mes, ano):
# #     try:
# #         df = pd.read_csv(caminho_balancete + '.csv', encoding='latin1', sep=';')
# #         print("Dados da primeira coluna e seus tipos:")
# #         for value in df.iloc[:, 0]:
# #             print(f"Valor: {value}, Tipo: {type(value)}")
# #         if '259' in df.iloc[:, 0].astype(str).values:
# #             linha_259 = df[df.iloc[:, 0].astype(str) == '259'].index[0]
# #             df = df.iloc[:linha_259 + 1, :]
# #             df.to_csv(caminho_balancete + '_processado.csv', index=False, sep=';', encoding='latin1')
# #             print(f"Arquivo processado salvo em: {caminho_balancete}_processado.csv")
# #             with open(caminho_balancete + '_processado.csv', 'r', encoding='latin1') as file:
# #                 data = file.readlines()[1:]
# #                 text_area.delete(1.0, tk.END)
# #                 text_area.insert(tk.END, ''.join(data))
# #             colar_dados_no_excel(df, codigo_empresa, mes, ano)
# #         else:
# #             print("String '259' não encontrada na primeira coluna.")
# #     except Exception as e:
# #         print(f"Erro ao processar o balancete: {e}")

# # # Função para colar os dados no arquivo Excel
# # def colar_dados_no_excel(df, codigo_empresa, mes, ano):
# #     try:
# #         caminho_excel_original = r'C:\Users\contabil18\Documents\projectRD\CONCILIACA_EMPRESA_XX_XXXX.xlsx'
# #         caminho_excel_novo = f'C:\\Users\\contabil18\\Documents\\projectRD\\CONCILIACA_{codigo_empresa}_{mes}_{ano}.xlsx'
# #         shutil.copyfile(caminho_excel_original, caminho_excel_novo)
# #         workbook = load_workbook(caminho_excel_novo)
# #         sheet = workbook.active
# #         with tempfile.NamedTemporaryFile(delete=False, mode='w', encoding='latin1', suffix='.txt') as temp_file:
# #             temp_file_path = temp_file.name
# #             df.to_csv(temp_file_path, index=False, sep='\t', encoding='latin1')
# #         os.startfile(temp_file_path)
# #         time.sleep(2)
# #         pyautogui.hotkey('ctrl', 'a')
# #         pyautogui.hotkey('ctrl', 'c')
# #         time.sleep(2)
# #         pyautogui.hotkey('alt', 'f4')
# #         os.startfile(caminho_excel_novo)
# #         time.sleep(5)
# #         pyautogui.hotkey('ctrl', 'g')
# #         pyautogui.write('A1')
# #         pyautogui.press('enter')
# #         time.sleep(1)
# #         pyautogui.hotkey('ctrl', 'v')
# #         time.sleep(2)
# #         pyautogui.hotkey('ctrl', 'u')
# #         pyautogui.hotkey('ctrl', 's')
# #         print(f"Dados colados no arquivo Excel: {caminho_excel_novo}")
# #     except Exception as e:
# #         print(f"Erro ao colar dados no Excel: {e}")

# # # Função para copiar o conteúdo da área de texto para a área de transferência
# # def copiar_para_area_de_transferencia():
# #     root.clipboard_clear()
# #     root.clipboard_append(text_area.get(1.0, tk.END))
# #     root.update()
# #     messagebox.showinfo("Informação", "Conteúdo copiado para a área de transferência.")

# # # Função para iniciar o processo com o código da empresa fornecido
# # def iniciar_processo():
# #     codigo_empresa = entry_codigo_empresa.get()
# #     mes = entry_mes.get()
# #     ano = entry_ano.get()
# #     if codigo_empresa and mes and ano:
# #         abrir_aplicativo()
# #         encontrar_texto()
# #         processar_empresa(codigo_empresa, mes, ano)
# #     else:
# #         messagebox.showwarning("Aviso", "Por favor, insira o código da empresa, mês e ano.")



#     try:
#         icon_location = pyautogui.locateOnScreen(r'C:\Users\contabil18\Documents\projectRD\icone.png')
#         if icon_location:
#             icon_center = pyautogui.center(icon_location)
#             pyautogui.moveTo(icon_center)
#             time.sleep(2)
#             pyautogui.click(button='left')
#             time.sleep(2)
#             pyautogui.moveTo(707, 253)
#             pyautogui.click(button='left')
#             time.sleep(10)
#             pyautogui.press('enter')
#             time.sleep(2)
#             caminho_balancete = f'C:\\Users\\contabil18\\Documents\\projectRD\\balancetes\\balancete_{codigo_empresa_str}'
#             pyautogui.write(caminho_balancete)
#             pyautogui.press('enter')
#             time.sleep(2)
#             processar_balancete(caminho_balancete, codigo_empresa_str, mes, ano)
#         else:
#             print("Ícone não encontrado na tela.")
#     except pyautogui.ImageNotFoundException:
#         print("Imagem não encontrada na tela.")

# # Função para processar o balancete baixado
# def processar_balancete(caminho_balancete, codigo_empresa, mes, ano):
#     try:
#         df = pd.read_csv(caminho_balancete + '.csv', encoding='latin1', sep=';')
#         print("Dados da primeira coluna e seus tipos:")
#         for value in df.iloc[:, 0]:
#             print(f"Valor: {value}, Tipo: {type(value)}")
#         if '259' in df.iloc[:, 0].astype(str).values:
#             linha_259 = df[df.iloc[:, 0].astype(str) == '259'].index[0]
#             df = df.iloc[:linha_259 + 1, :]
#             df.to_csv(caminho_balancete + '_processado.csv', index=False, sep=';', encoding='latin1')
#             print(f"Arquivo processado salvo em: {caminho_balancete}_processado.csv")
#             with open(caminho_balancete + '_processado.csv', 'r', encoding='latin1') as file:
#                 data = file.readlines()[1:]
#                 text_area.delete(1.0, tk.END)
#                 text_area.insert(tk.END, ''.join(data))
#             colar_dados_no_excel(df, codigo_empresa, mes, ano)
#         else:
#             print("String '259' não encontrada na primeira coluna.")
#     except Exception as e:
#         print(f"Erro ao processar o balancete: {e}")

# # Função para colar os dados no arquivo Excel
# def colar_dados_no_excel(df, codigo_empresa, mes, ano):
#     try:
#         caminho_excel_original = r'C:\Users\contabil18\Documents\projectRD\CONCILIACA_EMPRESA_XX_XXXX.xlsx'
#         caminho_excel_novo = f'C:\\Users\\contabil18\\Documents\\projectRD\\CONCILIACA_{codigo_empresa}_{mes}_{ano}.xlsx'
#         shutil.copyfile(caminho_excel_original, caminho_excel_novo)
#         workbook = load_workbook(caminho_excel_novo)
#         sheet = workbook.active
#         with tempfile.NamedTemporaryFile(delete=False, mode='w', encoding='latin1', suffix='.txt') as temp_file:
#             temp_file_path = temp_file.name
#             df.to_csv(temp_file_path, index=False, sep='\t', encoding='latin1')
#         os.startfile(temp_file_path)
#         time.sleep(2)
#         pyautogui.hotkey('ctrl', 'a')
#         pyautogui.hotkey('ctrl', 'c')
#         time.sleep(2)
#         pyautogui.hotkey('alt', 'f4')
#         os.startfile(caminho_excel_novo)
#         time.sleep(5)
#         pyautogui.hotkey('ctrl', 'g')
#         pyautogui.write('A1')
#         pyautogui.press('enter')
#         time.sleep(1)
#         pyautogui.hotkey('ctrl', 'v')
#         time.sleep(2)
#         pyautogui.hotkey('ctrl', 'u')
#         pyautogui.hotkey('ctrl', 's')
#         print(f"Dados colados no arquivo Excel: {caminho_excel_novo}")
#     except Exception as e:
#         print(f"Erro ao colar dados no Excel: {e}")

# # Função para copiar o conteúdo da área de texto para a área de transferência
# def copiar_para_area_de_transferencia():
#     root.clipboard_clear()
#     root.clipboard_append(text_area.get(1.0, tk.END))
#     root.update()
#     messagebox.showinfo("Informação", "Conteúdo copiado para a área de transferência.")

# # Função para iniciar o processo com o código da empresa fornecido
# def iniciar_processo():
#     codigo_empresa = entry_codigo_empresa.get()
#     mes = entry_mes.get()
#     ano = entry_ano.get()
#     if codigo_empresa and mes and ano:
#         abrir_aplicativo()
#         encontrar_texto()
#         processar_empresa(codigo_empresa, mes, ano)
#     else:
#         messagebox.showwarning("Aviso", "Por favor, insira o código da empresa, mês e ano.")

# # Configuração da interface gráfica
# root = tk.Tk()
# root.title("Processamento de Empresa")

# tk.Label(root, text="Código da Empresa:").pack(pady=10)
# entry_codigo_empresa = tk.Entry(root)
# entry_codigo_empresa.pack(pady=5)

# tk.Label(root, text="Mês:").pack(pady=10)
# entry_mes = tk.Entry(root)
# entry_mes.pack(pady=5)

# tk.Label(root, text="Ano:").pack(pady=10)
# entry_ano = tk.Entry(root)
# entry_ano.pack(pady=5)

# tk.Button(root, text="Iniciar Processo", command=iniciar_processo).pack(pady=20)

# text_area = scrolledtext.ScrolledText(root, width=80, height=20)
# text_area.pack(pady=10)

# tk.Button(root, text="Copiar para Área de Transferência", command=copiar_para_area_de_transferencia).pack(pady=10)

# root.mainloop()