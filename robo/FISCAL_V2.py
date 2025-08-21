from bs4 import BeautifulSoup
import openpyxl
import pyautogui
import time
import os
import pygetwindow as gw
import csv
import customtkinter as ctk
import tkinter as tk
import re
import subprocess
import sys
import glob
import calendar

pyautogui.FAILSAFE = False

#company_code = sys.argv[1]
#month_year = sys.argv[2]
#company_name = sys.argv[3]

def definir_competencia():
    def on_ok():
        nonlocal competencia, codigo_empresa
        competencia = entry_comp.get()
        codigo_empresa = entry_cod.get()
        dialog.destroy()
        root.quit()

    def on_timeout():
        nonlocal timeout
        timeout = True
        dialog.destroy()
        root.quit()

    # Lê empresas do CSV
    empresas = []
    try:
        with open(r"C:\projeto\empresas.csv", newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                empresas.append({'codigo': row['codigo'], 'nome': row['nome']})
    except Exception as e:
        empresas = []

    competencia = None
    codigo_empresa = None
    timeout = False

    root = tk.Tk()
    root.withdraw()

    dialog = tk.Toplevel(root)
    dialog.title("Definir Competência e Empresa")
    dialog.geometry("500x430")
    dialog.attributes('-topmost', True)

    label1 = tk.Label(dialog, text="Digite a competência no formato mmaaaa (ex: 052025):")
    label1.pack(pady=5)
    entry_comp = tk.Entry(dialog)
    entry_comp.pack(pady=5)
    entry_comp.focus()

    label2 = tk.Label(dialog, text="Digite o código da empresa:")
    label2.pack(pady=5)
    entry_cod = tk.Entry(dialog)
    entry_cod.pack(pady=5)

    label_search = tk.Label(dialog, text="Buscar empresa pelo nome:")
    label_search.pack(pady=5)
    entry_search = tk.Entry(dialog)
    entry_search.pack(pady=5)

    label3 = tk.Label(dialog, text="Códigos e empresas disponíveis:")
    label3.pack(pady=5)
    listbox = tk.Listbox(dialog, height=10, width=60)
    listbox.pack(pady=5, fill=tk.BOTH, expand=True)

    def update_listbox(filter_text=""):
        listbox.delete(0, tk.END)
        filtered = [
            emp for emp in empresas
            if filter_text.lower() in emp['nome'].lower()
        ] if filter_text else empresas
        for emp in filtered:
            listbox.insert(tk.END, f"{emp['codigo']} - {emp['nome']}")

    def on_listbox_select(event):
        selection = listbox.curselection()
        if selection:
            valor = listbox.get(selection[0])
            codigo = valor.split(" - ")[0]
            entry_cod.delete(0, tk.END)
            entry_cod.insert(0, codigo)

    def on_search_change(event):
        filter_text = entry_search.get()
        update_listbox(filter_text)

    listbox.bind('<<ListboxSelect>>', on_listbox_select)
    entry_search.bind('<KeyRelease>', on_search_change)

    update_listbox()

    # Botão OK (Confirmar escolha)
    btn = tk.Button(dialog, text="OK (Confirmar escolha)", command=on_ok, bg="#4CAF50", fg="white", font=("Arial", 12, "bold"))
    btn.pack(pady=15)

    dialog.after(20000, on_timeout)
    root.mainloop()

    if timeout:
        return None, None

    # Validação simples
    if (
        competencia and len(competencia) == 6 and competencia[:2].isdigit() and competencia[2:].isdigit()
        and codigo_empresa and codigo_empresa.isdigit()
    ):
        mes = int(competencia[:2])
        if 1 <= mes <= 12:
            return competencia, codigo_empresa
    return None, None

competencia, codigo_empresa = definir_competencia()
if competencia is None or codigo_empresa is None:
    print("A competência ou o código da empresa não foi definido ou o tempo expirou. Encerrando o programa.")
    sys.exit(0)

print(f"Competência escolhida: {competencia}")
print(f"Código da empresa escolhido: {codigo_empresa}")
time.sleep(5)
sys.stdout.reconfigure(encoding='utf-8')

# Busca o nome da empresa no CSV
company_name = None
try:
    with open(r"C:\projeto\empresas.csv", newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile, delimiter=';')
        for row in reader:
            if row['codigo'] == codigo_empresa:
                company_name = row['nome']
                break
except Exception as e:
    print(f"Erro ao ler arquivo de empresas: {e}")
    sys.exit(0)

if company_name is None:
    print(f"Empresa com código {codigo_empresa} não encontrada no arquivo CSV.")
    sys.exit(0)

company_code = codigo_empresa
month_year = competencia
#####################
# Função para obter o último dia do mês
def get_last_day_of_month(month_year):
    month = int(month_year[:2])
    year = int(month_year[2:])
    return calendar.monthrange(year, month)[1]

# Formatar as datas
month_year_formated = f'{month_year[0:2]}-{month_year[2:]}'
last_day = get_last_day_of_month(month_year)
month_year_end = f'{month_year[0:2]:0>2}-{month_year[2:]}'

caminho_html_csll_presumido_recolher = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório de apuração IRLP -  Sequência 22 - Ordem 9 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm"
csll_presumido_recolher = None
try:
    with open(caminho_html_csll_presumido_recolher, 'r', encoding='utf-8') as file:
        content = file.read()
        
    soup = BeautifulSoup(content, 'html.parser')
    results = soup.find_all('td', string=lambda text: text and "CSLL ACUMULADO TRIMESTRE" in text)

    total_temp = 0
    value_found = False
    for result in results:
        row = result.find_parent('tr')
        last_value = row.find_all('td')[-1].get_text(strip=True)
        try:
            total_temp += float(last_value.replace('.', '').replace(',', '.'))
            value_found = True
        except ValueError:
            continue
        
    if value_found:
        csll_presumido_recolher = round(float(total_temp), 2)
        if csll_presumido_recolher == 0:
            csll_presumido_recolher = None
        else:
            print(f"csll_presumido_recolher: {csll_presumido_recolher:.2f}".replace('.', ','))
    else:
        print("csll_presumido_recolher: Valor não encontrado no HTML")
        
except FileNotFoundError:
    print(f"Erro: O arquivo {caminho_html_csll_presumido_recolher} não existe. csll_presumido_recolher definido como None.")
    csll_presumido_recolher = None

#####################


caminho_html_irpj_presumido_recolher = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório de apuração IRLP -  Sequência 22 - Ordem 9 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm"
irpj_a_recolher_presumido = None
try:
    with open(caminho_html_irpj_presumido_recolher, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')
    results = soup.find_all('td', string=lambda text: text and "IRPJ ACUMULADO TRIMESTRE" in text)

    total_temp = 0
    value_found = False
    for result in results:
        row = result.find_parent('tr')
        last_value = row.find_all('td')[-1].get_text(strip=True)
        try:
            total_temp += float(last_value.replace('.', '').replace(',', '.'))
            value_found = True
        except ValueError:
            continue

    if value_found:
        irpj_a_recolher_presumido = round(float(total_temp), 2)
        if irpj_a_recolher_presumido == 0:
            irpj_a_recolher_presumido = None
        else:
            print(f"irpj_presumido_recolher: {irpj_a_recolher_presumido:.2f}".replace('.', ','))
    else:
        print("irpj_presumido_recolher: Valor não encontrado no HTML")
        
except FileNotFoundError:
    print(f"Erro: O arquivo {caminho_html_irpj_presumido_recolher} não existe. irpj_a_recolher_presumido definido como None.")
    irpj_a_recolher_presumido = None

#####################

caminho_html_icms_cp_recolher = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório apuração DimeSC -  Sequência 22 - Ordem 1 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm"
total_icms_cp_recolher = None
try:
    with open(caminho_html_icms_cp_recolher, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')
    results = soup.find_all('td', string=lambda text: text and "(=) Imposto a recolher pela utilização do crédito presumido" in text)

    total_temp = 0
    value_found = False
    for result in results:
        row = result.find_parent('tr')
        last_value = row.find_all('td')[-1].get_text(strip=True)
        try:
            total_temp += float(last_value.replace('.', '').replace(',', '.'))
            value_found = True
        except ValueError:
            continue

    if value_found:
        total_icms_cp_recolher = round(float(total_temp), 2)
        if total_icms_cp_recolher == 0:
            total_icms_cp_recolher = None
        else:
            print(f"ICMS_cp_recolher: {total_icms_cp_recolher:.2f}".replace('.', ','))
    else:
        print("ICMS_cp_recolher: Valor não encontrado no HTML")
        
except FileNotFoundError:
    print(f"Erro: O arquivo {caminho_html_icms_cp_recolher} não existe. total_icms_cp_recolher definido como None.")
    total_icms_cp_recolher = None

#####################

caminho_issqn_a_recolher = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Conferência serviços -  Sequência 22 - Ordem 10 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm"
issqn_a_recolher = None
try:
    with open(caminho_issqn_a_recolher, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')
    result = soup.find('td', string="TOTAL GERAL:")

    if result:
        row = result.find_parent('tr')
        last_value = row.find_all('td')[-2].get_text(strip=True)
        print(f"iss_qn_a_recolher: {last_value}")  # Adiciona uma instrução de depuração aqui
        try:
            # Remove caracteres não numéricos, exceto vírgula e ponto
            last_value_cleaned = re.sub(r'[^\d,]', '', last_value)
            issqn_a_recolher = float(last_value_cleaned.replace('.', '').replace(',', '.'))
            if issqn_a_recolher == 0:
                issqn_a_recolher = None
        except ValueError:
            print(f"Erro ao converter o valor: {last_value}")  # Adiciona uma instrução de depuração aqui
            issqn_a_recolher = None
    else:
        print("issqn_a_recolher: Valor 'TOTAL GERAL:' não encontrado no HTML")
        
except FileNotFoundError:
    print(f"Erro: O arquivo {caminho_issqn_a_recolher} não existe. issqn_a_recolher definido como None.")
    issqn_a_recolher = None

#####################

caminho_sn_a_recolher = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Demonstrativo simples nacional -  Sequência 22 - Ordem 8 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm"
sn_a_recolher = None
try:
    with open(caminho_sn_a_recolher, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')
    results = soup.find_all('td', string=lambda text: text and "Valor do Simples a Pagar" in text)

    if not results:
        print("Nenhum resultado encontrado para 'Valor do Simples a Pagar'")
        sn_a_recolher = None
    else:
        total_temp = 0
        value_found = False
        for result in results:
            row = result.find_parent('tr')
            last_value = row.find_all('td')[-1].get_text(strip=True)
            print(f"Valor encontrado: {last_value}")  # Adiciona uma instrução de depuração aqui
            try:
                # Remove caracteres não numéricos, exceto vírgula e ponto
                last_value_cleaned = re.sub(r'[^\d,]', '', last_value)
                total_temp += float(last_value_cleaned.replace('.', '').replace(',', '.'))
                value_found = True
            except ValueError:
                print(f"Erro ao converter o valor: {last_value}")  # Adiciona uma instrução de depuração aqui
                continue

        if value_found:
            sn_a_recolher = round(float(total_temp), 2)
            if sn_a_recolher == 0:
                sn_a_recolher = None
            else:
                print(f"sn_recolher: {sn_a_recolher:.2f}".replace('.', ','))
        else:
            print("sn_a_recolher: Valor não encontrado no HTML")
            
except FileNotFoundError:
    print(f"Erro: O arquivo {caminho_sn_a_recolher} não existe. sn_a_recolher definido como None.")
    sn_a_recolher = None
except Exception as e:
    print(f"Erro ao abrir o arquivo {caminho_sn_a_recolher}: {e}")
    sn_a_recolher = None

#####################

caminho_html_icms_fundos = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório apuração DimeSC -  Sequência 22 - Ordem 1 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm"
total_icms_fundos = None
fundos_a_recolher = None
try:
    with open(caminho_html_icms_fundos, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')
        
    # Procurar pelo valor do FUNDO SOCIAL A RECOLHER
    fundo_social_result = soup.find('td', string="(=) FUNDO SOCIAL A RECOLHER")
    if fundo_social_result:
        row = fundo_social_result.find_parent('tr')
        fundo_social_value_td = row.find_all('td', class_='s9')[-1]
        fundo_social_value = fundo_social_value_td.get_text(strip=True)
        fundo_social_value_numbers = ''.join(filter(str.isdigit, fundo_social_value.replace(',', '')))
        fundo_social_a_recolher = float(f"{int(fundo_social_value_numbers[:-2])}.{fundo_social_value_numbers[-2:]}")
    else:
        fundo_social_a_recolher = None

    # Procurar pelo valor do FUMDES A RECOLHER
    fumdes_result = soup.find('td', string="(=) FUMDES A RECOLHER")
    if fumdes_result:
        row = fumdes_result.find_parent('tr')
        fumdes_value_td = row.find_all('td', class_='s9')[-1]
        fumdes_value = fumdes_value_td.get_text(strip=True)
        fumdes_value_numbers = ''.join(filter(str.isdigit, fumdes_value.replace(',', '')))
        fumdes_a_recolher = float(f"{int(fumdes_value_numbers[:-2])}.{fumdes_value_numbers[-2:]}")
    else:
        fumdes_a_recolher = None

    # Somar os valores apenas se ambos existirem
    if fundo_social_a_recolher is not None and fumdes_a_recolher is not None:
        fundos_a_recolher = fundo_social_a_recolher + fumdes_a_recolher
    elif fundo_social_a_recolher is not None:
        fundos_a_recolher = fundo_social_a_recolher
    elif fumdes_a_recolher is not None:
        fundos_a_recolher = fumdes_a_recolher
    else:
        fundos_a_recolher = None
        print("fundos_a_recolher: Valores não encontrados no HTML")

    if fundos_a_recolher is not None:
        if fundos_a_recolher == 0:
            fundos_a_recolher = None
        else:
            formatted_value = f"{fundos_a_recolher:,.2f}".replace('.', ',').replace(',', '', 1)
            print(f"fundos_a_recolher: {formatted_value}")

except FileNotFoundError:
    print(f"Erro: O arquivo {caminho_html_icms_fundos} não existe. fundos_a_recolher definido como None.")
    fundos_a_recolher = None
except Exception as e:
    print(f"Erro ao abrir o arquivo {caminho_html_icms_fundos}: {e}")
    fundos_a_recolher = None

#####################

caminho_html_icms_a_recolher = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório apuração DimeSC -  Sequência 22 - Ordem 1 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm"
total_icms_a_recolher = None
try:
    with open(caminho_html_icms_a_recolher, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')
    results = soup.find_all('td', string=lambda text: text and "(=)Imposto a recolher" in text)

    total_temp = 0
    value_found = False
    for result in results:
        row = result.find_parent('tr')
        last_value = row.find_all('td')[-1].get_text(strip=True)
        try:
            total_temp += float(last_value.replace('.', '').replace(',', '.'))
            value_found = True
        except ValueError:
            continue

    if value_found:
        total_icms_a_recolher = round(float(total_temp), 2)
        if total_icms_a_recolher == 0:
            total_icms_a_recolher = None
        else:
            print(f"icms_a_recolher: {total_icms_a_recolher}")
    else:
        print("total_icms_a_recolher: Valor não encontrado no HTML")
        
except FileNotFoundError:
    print(f"Erro: O arquivo {caminho_html_icms_a_recolher} não existe. total_icms_a_recolher definido como None.")
    total_icms_a_recolher = None

#####################

caminho_html_ipi_recolher = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório apuração IPI -  Sequência 22 - Ordem 2 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm"
total_ipi_recolher = None

# Lista de possíveis variações do nome do arquivo IPI
possible_ipi_files = [
    f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório apuração IPI -  Sequência 22 - Ordem 2 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm",
    f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório apuração IPI - Sequência 22 - Ordem 2 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm",
    f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório apuração IPI -   Sequência 22 - Ordem 2 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm"
]

# Tenta encontrar o arquivo correto
found_file = None
for filepath in possible_ipi_files:
    if os.path.exists(filepath):
        found_file = filepath
        caminho_html_ipi_recolher = filepath
        break

if found_file is None:
    print("Aviso: Tentando encontrar arquivo IPI com variações de nome...")
    import glob
    pattern = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório apuração IPI*Sequência 22*Ordem 2*{month_year_formated}*.htm"
    matching_files = glob.glob(pattern)
    if matching_files:
        caminho_html_ipi_recolher = matching_files[0]
        found_file = matching_files[0]
        print(f"Arquivo IPI encontrado: {os.path.basename(found_file)}")

try:
    if found_file is None:
        raise FileNotFoundError(f"Nenhum arquivo IPI encontrado para empresa {company_code}")
        
    with open(caminho_html_ipi_recolher, 'r', encoding='utf-8') as file:
        content = file.read()
    
    print(f"Processando arquivo IPI: {caminho_html_ipi_recolher}")
    
    # Verifica se o arquivo contém texto relacionado ao IPI
    if 'IPI' not in content:
        print("Aviso: O arquivo não contém texto 'IPI'")
    if 'recolher' not in content:
        print("Aviso: O arquivo não contém texto 'recolher'")
    
    soup = BeautifulSoup(content, 'html.parser')
    
    # Múltiplas abordagens para encontrar o valor do IPI
    
    # Debug: Mostra se encontrou células relacionadas ao IPI
    debug_cells = soup.find_all('td', string=lambda text: text and 'IPI' in text)
    if debug_cells:
        print(f"Debug: Encontradas {len(debug_cells)} células contendo 'IPI':")
        for i, cell in enumerate(debug_cells[:5]):  # Mostra até 5 células
            print(f"  {i+1}: '{cell.get_text(strip=True)}'")
    
    # Abordagem 1: Busca pela estrutura específica mencionada
    text_cell = soup.find('td', {'colspan': '7', 'class': 's6'}, string=lambda text: text and "Valor do saldo devedor do IPI a recolher" in text)
    
    if text_cell:
        # Busca a linha que contém o texto
        row = text_cell.find_parent('tr')
        
        # Busca todas as células na linha
        cells = row.find_all('td')
        
        # Procura por valores monetários nas células
        for cell in cells:
            cell_text = cell.get_text(strip=True)
            # Verifica se parece com um valor monetário (contém vírgula ou é um número grande)
            if ',' in cell_text or ('.' in cell_text and len(cell_text.replace('.', '').replace(',', '')) > 2):
                try:
                    test_value = float(cell_text.replace('.', '').replace(',', '.'))
                    if test_value >= 100:  # Valor monetário significativo
                        total_ipi_recolher = test_value
                        if total_ipi_recolher == 0:
                            total_ipi_recolher = None
                        else:
                            print(f"ipi_recolher (método 1): {total_ipi_recolher:.2f}".replace('.', ','))
                        break
                except ValueError:
                    continue
    
    # Abordagem 2: Se não encontrou pelo método 1, tenta busca mais ampla
    if total_ipi_recolher is None:
        # Procura por qualquer texto que contenha "IPI a recolher"
        ipi_cells = soup.find_all('td', string=lambda text: text and "IPI a recolher" in text)
        
        for cell in ipi_cells:
            row = cell.find_parent('tr')
            cells = row.find_all('td')
            
            # Procura por valores monetários nesta linha
            for value_cell in cells:
                value_text = value_cell.get_text(strip=True)
                if ',' in value_text or ('.' in value_text and len(value_text.replace('.', '').replace(',', '')) > 2):
                    try:
                        test_value = float(value_text.replace('.', '').replace(',', '.'))
                        if test_value >= 100:  # Valor monetário significativo
                            total_ipi_recolher = test_value
                            if total_ipi_recolher == 0:
                                total_ipi_recolher = None
                            else:
                                print(f"ipi_recolher (método 2): {total_ipi_recolher:.2f}".replace('.', ','))
                            break
                    except ValueError:
                        continue
            if total_ipi_recolher is not None:
                break
    
    # Abordagem 3: Busca por padrões de valores monetários perto do texto IPI
    if total_ipi_recolher is None:
        # Busca por todas as células que contêm valores monetários significativos
        all_cells = soup.find_all('td')
        monetary_values = []
        
        for cell in all_cells:
            cell_text = cell.get_text(strip=True)
            if ',' in cell_text and '.' in cell_text:
                try:
                    test_value = float(cell_text.replace('.', '').replace(',', '.'))
                    if test_value >= 100:  # Valor monetário significativo
                        # Verifica se há menção de IPI nas células próximas
                        row = cell.find_parent('tr')
                        row_text = row.get_text().lower()
                        if 'ipi' in row_text and 'recolher' in row_text:
                            monetary_values.append(test_value)
                except ValueError:
                    continue
        
        if monetary_values:
            # Pega o maior valor (mais provável de ser o total)
            total_ipi_recolher = max(monetary_values)
            if total_ipi_recolher == 0:
                total_ipi_recolher = None
            else:
                print(f"ipi_recolher (método 3): {total_ipi_recolher:.2f}".replace('.', ','))
    
    if total_ipi_recolher is None:
        print("total_ipi_recolher: Valor não encontrado no HTML usando todas as abordagens")
        
except FileNotFoundError:
    print(f"Erro: O arquivo {caminho_html_ipi_recolher} não existe. total_ipi_recolher definido como None.")
    total_ipi_recolher = None
except Exception as e:
    print(f"Erro ao processar arquivo IPI: {e}")
    total_ipi_recolher = None


#####################

caminho_html_cofins_recolher = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório de apuração de PISpasep e Cofins -  Sequência 22 - Ordem 4 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm"
total_cofins_recolher = None
try:
    with open(caminho_html_cofins_recolher, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')
    results = soup.find_all('td', string=lambda text: text and "Valor total da contribuição a recolher/pagar no período (08 + 12)" in text)

    total_temp = 0
    value_found = False
    for result in results:
        row = result.find_parent('tr')
        last_value = row.find_all('td')[-1].get_text(strip=True)
        try:
            total_temp += float(last_value.replace('.', '').replace(',', '.'))
            value_found = True
        except ValueError:
            continue

    if value_found:
        total_cofins_recolher = round(float(total_temp), 2)
        if total_cofins_recolher == 0:
            total_cofins_recolher = None
        else:
            print(f"cofins_recolher: {total_cofins_recolher:.2f}".replace('.', ','))
    else:
        print("total_cofins_recolher: Valor não encontrado no HTML")
        
except FileNotFoundError:
    print(f"Erro: O arquivo {caminho_html_cofins_recolher} não existe. total_cofins_recolher definido como None.")
    total_cofins_recolher = None

#####################

caminho_html_pis_recolher = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório de apuração de PISpasep e Cofins -  Sequência 22 - Ordem 4 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm"
total_pis_recolher = None
try:
    with open(caminho_html_pis_recolher, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')
    results = soup.find_all('td', string=lambda text: text and "Valor total da contribuição a recolher/pagar no período (08 + 12)" in text)

    total_temp = 0
    value_found = False
    for result in results:
        row = result.find_parent('tr')
        last_value = row.find_all('td')[-2].get_text(strip=True)
        try:
            total_temp += float(last_value.replace('.', '').replace(',', '.'))
            value_found = True
        except ValueError:
            continue

    if value_found:
        total_pis_recolher = round(float(total_temp), 2)
        if total_pis_recolher == 0:
            total_pis_recolher = None
        else:
            print(f"PIS_recolher: {total_pis_recolher:.2f}".replace('.', ','))
    else:
        print("total_pis_recolher: Valor não encontrado no HTML")
        
except FileNotFoundError:
    print(f"Erro: O arquivo {caminho_html_pis_recolher} não existe. total_pis_recolher definido como None.")
    total_pis_recolher = None

#####################

try:
    with open(f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório de retenções -  Sequência 22 - Ordem 5 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm", 'r', encoding='utf-8') as file:
        html_content = file.read()

    # Parsear o HTML
    soup = BeautifulSoup(html_content, 'html.parser')

    linha = soup.find_all('td', class_='s8')  # Busca todas as células com classe 's8'

    # Inicializa linha_completa como None
    linha_completa = None

    # Filtrar com base no texto 'Total Geral'
    for td in linha:
        if 'Total Geral' in td.get_text():
            linha_completa = td.find_parent('tr')  # Encontra a linha (<tr>) associada à célula

    # Verifica se linha_completa foi definida antes de usá-la
    if linha_completa:
        # Extrair valores dinâmicos, ignorando a primeira observação e formatando os valores
        valores = []
        for td in linha_completa.find_all('td')[1:]:
            try:
                valor = float(td.get_text(strip=True).replace('.', '').replace(',', '.'))
                valores.append(round(valor, 2))
            except ValueError:
                continue

        # Verificar se temos valores suficientes
        if len(valores) >= 6:
            # Atribuir valores às variáveis
            Pis_retido = valores[1] if len(valores) > 1 else None
            Cofins_retido = valores[2] if len(valores) > 2 else None
            csll_retido = valores[3] if len(valores) > 3 else None
            irrf_retido = valores[4] if len(valores) > 4 else None
            iss_retido = valores[5] if len(valores) > 5 else None
            inss_retido = valores[6] if len(valores) > 6 else None

            # Calcular o total de CSRF retido apenas se os valores existirem
            if Pis_retido is not None and Cofins_retido is not None and csll_retido is not None:
                csrf_retido = Pis_retido + Cofins_retido + csll_retido
            else:
                csrf_retido = None

            # Exibir valores retidos
            if irrf_retido is not None:
                print(f"IRRF Retido: {irrf_retido:.2f}")
            else:
                print("IRRF Retido: Valor não encontrado no HTML")
            
            if iss_retido is not None:
                print(f"ISS Retido: {iss_retido:.2f}")
            else:
                print("ISS Retido: Valor não encontrado no HTML")
            
            if inss_retido is not None:
                print(f"INSS Retido: {inss_retido:.2f}")
            else:
                print("INSS Retido: Valor não encontrado no HTML")
            
            if csrf_retido is not None:
                print(f"CSRF Retido: {csrf_retido:.2f}")
            else:
                print("CSRF Retido: Valor não encontrado no HTML")
        else:
            print("Valores de retenção não encontrados no HTML - dados insuficientes")
            Pis_retido = Cofins_retido = csll_retido = irrf_retido = iss_retido = inss_retido = csrf_retido = None
    else:
        print("Linha 'Total Geral' não encontrada no HTML de retenções")
        Pis_retido = Cofins_retido = csll_retido = irrf_retido = iss_retido = inss_retido = csrf_retido = None

except FileNotFoundError:
    print(f"Erro: Arquivo de retenções não existe. Valores de retenção definidos como None.")
    Pis_retido = Cofins_retido = csll_retido = irrf_retido = iss_retido = inss_retido = csrf_retido = None
except Exception as e:
    print(f"Erro ao processar arquivo de retenções: {e}")
    Pis_retido = Cofins_retido = csll_retido = irrf_retido = iss_retido = inss_retido = csrf_retido = None

#####################

caminho_html_icms_cp_recup = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório apuração DimeSC -  Sequência 22 - Ordem 1 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm"
total_icms_cp_recup = None
try:
    with open(caminho_html_icms_cp_recup, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')
    results = soup.find_all('td', string=lambda text: text and "(=) Saldo credor das antecipações para o mês seguinte" in text)

    total_temp = 0
    value_found = False
    for result in results:
        row = result.find_parent('tr')
        last_value = row.find_all('td')[-2].get_text(strip=True)
        try:
            total_temp += float(last_value.replace('.', '').replace(',', '.'))
            value_found = True
        except ValueError:
            continue

    if value_found:
        total_icms_cp_recup = round(float(total_temp), 2)
        if total_icms_cp_recup == 0:
            total_icms_cp_recup = None
        else:
            print(f"ICMS_cp_recup: {total_icms_cp_recup:.2f}".replace('.', ','))
    else:
        print("total_icms_cp_recup: Valor não encontrado no HTML")
        
except FileNotFoundError:
    print(f"Erro: O arquivo {caminho_html_icms_cp_recup} não existe. total_icms_cp_recup definido como None.")
    total_icms_cp_recup = None

#####################

caminho_html_cofins_recup = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório de apuração de PISpasep e Cofins -  Sequência 22 - Ordem 4 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm"
total_cofins_recup = None
try:
    with open(caminho_html_cofins_recup, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')
    results = soup.find_all('td', string=lambda text: text and "Totais Cofins" in text)

    total_temp = 0
    value_found = False
    for result in results:
        row = result.find_parent('tr')
        last_value = row.find_all('td')[-1].get_text(strip=True)
        try:
            total_temp += float(last_value.replace('.', '').replace(',', '.'))
            value_found = True
        except ValueError:
            continue

    if value_found:
        total_cofins_recup = round(float(total_temp), 2)
        if total_cofins_recup == 0:
            total_cofins_recup = None
        else:
            print(f"COFINS_recup: {total_cofins_recup:.2f}".replace('.', ','))
    else:
        print("total_cofins_recup: Valor não encontrado no HTML")
        
except FileNotFoundError:
    print(f"Erro: O arquivo {caminho_html_cofins_recup} não existe. total_cofins_recup definido como None.")
    total_cofins_recup = None

#####################


caminho_html_pis_recup = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório de apuração de PISpasep e Cofins -  Sequência 22 - Ordem 4 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm"
total_pis_recup = None
try:
    with open(caminho_html_pis_recup, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')
    results = soup.find_all('td', string=lambda text: text and "Totais Pis/Pasep" in text)

    total_temp = 0
    value_found = False
    for result in results:
        row = result.find_parent('tr')
        last_value = row.find_all('td')[-1].get_text(strip=True)
        try:
            total_temp += float(last_value.replace('.', '').replace(',', '.'))
            value_found = True
        except ValueError:
            continue

    if value_found:
        total_pis_recup = round(float(total_temp), 2)
        if total_pis_recup == 0:
            total_pis_recup = None
        else:
            print(f"PIS_recup: {total_pis_recup:.2f}".replace('.', ','))
    else:
        print("total_pis_recup: Valor não encontrado no HTML")
        
except FileNotFoundError:
    print(f"Erro: O arquivo {caminho_html_pis_recup} não existe. total_pis_recup definido como None.")
    total_pis_recup = None

#####################

# Montar o nome do arquivo Excel corretamente substituindo espaços e barras por underline
nome_empresa_arquivo = company_name.replace(' ', '_').replace('/', '_').replace('\\', '_')
excel_path = f'C:\\projeto\\planilhas\\CONCILIACAO_{nome_empresa_arquivo}_{month_year}.xlsx'
wb = openpyxl.load_workbook(excel_path)
ws = wb.active
  
###################################  
    
numeros_procurados = [43, 48, 41, 42, 46, 47, 2654, 617, 185, 2707, 186, 197, 196, 198, 195, 199, 201, 202, 191, 190, 192, 203]

# Mapeamento de cell_a para os valores correspondentes
valor_map = {
    41: 'ICMS_recup',
    42: 'ipi_a_recup',
    46: 'total_pis_recup',
    47: 'total_cofins_recup',
    2654: 'total_icms_cp_recup',
    617: 'csrf_retido',
    185: 'irrf_retido',
    2707: 'inss_retido',
    186: 'iss_retido',
    197: 'total_pis_recolher',
    196: 'total_cofins_recolher',
    198: 'total_ipi_recolher',
    195: 'total_icms_a_recolher',
    199: 'fundos_a_recolher',
    202: 'sn_a_recolher',
    201: 'issqn_a_recolher',
    191: 'total_icms_cp_recolher',
    190: 'irpj_a_recolher_presumido',
    192: 'csll_presumido_recolher',
    48: 'total_csll_recup',
    43: 'total_irrf_recup',
    203: 'total_icms_st'
}

###################################

def extract_saldo_credor(html_path_ipi_a_recup):
    try:
        with open(f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório apuração IPI -  Sequência 22 - Ordem 2 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm", 'r', encoding='utf-8') as file:
            html_content = file.read()
            
        soup = BeautifulSoup(html_content, 'html.parser')
        saldo_credor_element = soup.find('td', string="SALDO CREDOR PERIODO SEGUINTE")

        if saldo_credor_element:
            ipi_a_recup = saldo_credor_element.find_next_sibling('td').get_text(strip=True)
            ipi_a_recup = float(ipi_a_recup.replace('.', '').replace(',', '.'))
            return ipi_a_recup
        else:
            print("ipi_a_recup: Variável 'SALDO CREDOR PERIODO SEGUINTE' não encontrada no HTML")
            return None
    except FileNotFoundError:
        print(f"Erro: Arquivo IPI não existe. ipi_a_recup definido como None.")
        return None
    except Exception as e:
        print(f"Erro ao processar arquivo IPI: {e}")
        return None
    
html_path_ipi_a_recup = f'C:\\projeto\\planilhas\\balancete\\CONCILIACAO_{company_code}_{month_year}.xlsx'
ipi_a_recup = extract_saldo_credor(html_path_ipi_a_recup)
if ipi_a_recup is not None:
    if ipi_a_recup == 0:
        ipi_a_recup = None
    else:
        print(f"IPI a recup: {ipi_a_recup:.2f}".replace('.', ','))
else:
    print("IPI a recup: Valor não encontrado no HTML")


# Restaura o bloco de leitura e impressão de ICMS_recup, CSLL_recup, IRRF_recup e PIS_recup antes do processamento da planilha.

########################

html_path_icms_recup = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório apuração DimeSC -  Sequência 22 - Ordem 1 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm"
ICMS_recup = None
try:
    with open(html_path_icms_recup, 'r', encoding='utf-8') as file:
        content = file.read()

    # Cria um objeto BeautifulSoup
    soup = BeautifulSoup(content, 'html.parser')

    # Busca a linha que começa com <td colspan="2" class="s9">190</td>
    result = soup.find('td', {'colspan': '2', 'class': 's9'}, string='190')

    # Exibe a linha inteira
    if result:
        row = result.find_parent('tr')
        observations = row.find_all('td')
        if len(observations) > 2:
            icms_text = observations[2].text
            icms_cleaned = icms_text.replace('.', '').replace(',', '.')
            ICMS_recup = float(icms_cleaned)
            if ICMS_recup == 0:
                ICMS_recup = None
            else:
                print(f"ICMS_recup: {ICMS_recup:.2f}".replace('.', ','))
        else:
            print("ICMS_recup: Dados insuficientes na linha encontrada")
    else:
        print("ICMS_recup: Linha com código '190' não encontrada no HTML")
        
except FileNotFoundError:
    print(f"Erro: O arquivo {html_path_icms_recup} não existe. ICMS_recup definido como None.")
    ICMS_recup = None
except Exception as e:
    print(f"Erro ao processar arquivo ICMS: {e}")
    ICMS_recup = None

caminho_html_csll_recup = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório de apuração de CSLL -  Sequência 22 - Ordem 4 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm"
total_csll_recup = None
try:
    with open(caminho_html_csll_recup, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')
    results = soup.find_all('td', string=lambda text: text and "Totais CSLL" in text)

    total_temp = 0
    value_found = False
    for result in results:
        row = result.find_parent('tr')
        last_value = row.find_all('td')[-1].get_text(strip=True)
        try:
            total_temp += float(last_value.replace('.', '').replace(',', '.'))
            value_found = True
        except ValueError:
            continue

    if value_found:
        total_csll_recup = round(float(total_temp), 2)
        if total_csll_recup == 0:
            total_csll_recup = None
        else:
            print(f"CSLL_recup: {total_csll_recup:.2f}".replace('.', ','))
    else:
        print("total_csll_recup: Valor não encontrado no HTML")
        
except FileNotFoundError:
    print(f"Erro: O arquivo {caminho_html_csll_recup} não existe. total_csll_recup definido como None.")
    total_csll_recup = None

caminho_html_irrf_recup = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório de apuração de IRRF -  Sequência 22 - Ordem 4 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm"
total_irrf_recup = None
try:
    with open(caminho_html_irrf_recup, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')
    results = soup.find_all('td', string=lambda text: text and "Totais IRRF" in text)

    total_temp = 0
    value_found = False
    for result in results:
        row = result.find_parent('tr')
        last_value = row.find_all('td')[-1].get_text(strip=True)
        try:
            total_temp += float(last_value.replace('.', '').replace(',', '.'))
            value_found = True
        except ValueError:
            continue

    if value_found:
        total_irrf_recup = round(float(total_temp), 2)
        if total_irrf_recup == 0:
            total_irrf_recup = None
        else:
            print(f"IRRF_recup: {total_irrf_recup:.2f}".replace('.', ','))
    else:
        print("total_irrf_recup: Valor não encontrado no HTML")
        
except FileNotFoundError:
    print(f"Erro: O arquivo {caminho_html_irrf_recup} não existe. total_irrf_recup definido como None.")
    total_irrf_recup = None

#####################

# Extração de ICMS ST
caminho_html_icms_st = f"C:\\relatorios_fiscal\\Empresa {company_code} - {company_name} - Relatório apuração DimeSC -  Sequência 22 - Ordem 1 01-{month_year_formated} {last_day:02d}-{month_year_formated}.htm"
total_icms_st = None
try:
    with open(caminho_html_icms_st, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')
    results = soup.find_all('td', string=lambda text: text and "ICMS ST a recolher" in text)

    total_temp = 0
    value_found = False
    for result in results:
        row = result.find_parent('tr')
        last_value = row.find_all('td')[-1].get_text(strip=True)
        try:
            total_temp += float(last_value.replace('.', '').replace(',', '.'))
            value_found = True
        except ValueError:
            continue

    if value_found:
        total_icms_st = round(float(total_temp), 2)
        if total_icms_st == 0:
            total_icms_st = None
        else:
            print(f"ICMS ST: {total_icms_st:.2f}".replace('.', ','))
    else:
        print("total_icms_st: Valor não encontrado no HTML")
        
except FileNotFoundError:
    print(f"Erro: O arquivo {caminho_html_icms_st} não existe. total_icms_st definido como None.")
    total_icms_st = None
except Exception as e:
    print(f"Erro ao processar arquivo ICMS ST: {e}")
    total_icms_st = None

#####################

# Extract the month from the user's input
input_month = month_year[:2]

for row in ws.iter_rows(min_row=2):
    cell_a = row[0].value  # Coluna A (índice 0)
    if cell_a in numeros_procurados:
        valor_coluna_f = row[5].value  # Coluna F (índice 5)
        valor_coluna_g = row[6].value  # Coluna G (índice 6)
        valor_coluna_h = row[7].value  # Coluna H (índice 7)
        # Converter valores para float se não forem None
        def to_float(val):
            if val is None:
                return None
            if isinstance(val, (int, float)):
                return float(val)
            try:
                # Se o valor estiver entre parênteses, extrai apenas o conteúdo interno
                val_str = str(val)
                match = re.match(r'^\((.*)\)$', val_str.strip())
                if match:
                    val_str = match.group(1)
                return float(val_str.replace('.', '').replace(',', '.'))
            except Exception:
                return None
        valor_coluna_f_float = to_float(valor_coluna_f)
        valor_coluna_g_float = to_float(valor_coluna_g)
        valor_coluna_h_float = to_float(valor_coluna_h)
        print(f"Número {cell_a} encontrado: F={valor_coluna_f_float}, G={valor_coluna_g_float}, H={valor_coluna_h_float}")
        # Comparar os valores e escrever "OK" ou "Verificar" na coluna I
        try:
            if cell_a in valor_map:
                valor_nome = valor_map[cell_a]
                try:
                    valor = globals()[valor_nome]
                    print(f"Comparando: {valor_nome} (valor={valor}) com F={valor_coluna_f_float}, G={valor_coluna_g_float}, H={valor_coluna_h_float}")
                    
                    # Define a tolerância de 10 centavos para todas as comparações
                    tolerancia = 0.10
                    
                    # Se o valor não foi encontrado no HTML (None), marcar como "Verificar"
                    if valor is None:
                        ws.cell(row=row[0].row, column=9, value="Verificar")
                        ws.cell(row=row[0].row, column=10, value=f"{valor_nome}: Valor não encontrado no HTML")
                        print(f"Linha {row[0].row}: Escreveu Verificar - Valor não encontrado no HTML")
                    elif cell_a == 190 and input_month in ['01', '02', '05', '08', '11']:
                        valor_comparacao = valor_coluna_g_float - valor_coluna_f_float if valor_coluna_g_float is not None and valor_coluna_f_float is not None else None
                        print(f"Comparando (G-F)={valor_comparacao} com {valor}")
                        ws.cell(row=row[0].row, column=10, value=f"IRPJ Presumido: (G-F)={valor_comparacao} vs {valor_nome}={valor}")
                        if valor_comparacao is not None and valor is not None and abs(valor_comparacao - valor) <= tolerancia:
                            ws.cell(row=row[0].row, column=9, value="OK")
                            print(f"Linha {row[0].row}: Escreveu OK (diferença: {abs(valor_comparacao - valor)})")
                        else:
                            ws.cell(row=row[0].row, column=9, value="Verificar")
                            diferenca = abs(valor_comparacao - valor) if valor_comparacao is not None and valor is not None else 'N/A'
                            print(f"Linha {row[0].row}: Escreveu Verificar (diferença: {diferenca})")
                    elif cell_a == 192 and input_month in ['01', '02', '05', '08', '11']:
                        valor_comparacao = valor_coluna_g_float - valor_coluna_f_float if valor_coluna_g_float is not None and valor_coluna_f_float is not None else None
                        print(f"Comparando (G-F)={valor_comparacao} com {valor}")
                        ws.cell(row=row[0].row, column=10, value=f"CSLL Presumido: (G-F)={valor_comparacao} vs {valor_nome}={valor}")
                        if valor_comparacao is not None and valor is not None and abs(valor_comparacao - valor) <= tolerancia:
                            ws.cell(row=row[0].row, column=9, value="OK")
                            print(f"Linha {row[0].row}: Escreveu OK (diferença: {abs(valor_comparacao - valor)})")
                        else:
                            ws.cell(row=row[0].row, column=9, value="Verificar")
                            diferenca = abs(valor_comparacao - valor) if valor_comparacao is not None and valor is not None else 'N/A'
                            print(f"Linha {row[0].row}: Escreveu Verificar (diferença: {diferenca})")
                    elif cell_a == 203:
                        ws.cell(row=row[0].row, column=10, value=f"ICMS ST: G={valor_coluna_g_float} vs {valor_nome}={valor}")
                        if valor_coluna_g_float is not None and valor is not None and abs(valor_coluna_g_float - valor) <= tolerancia:
                            ws.cell(row=row[0].row, column=9, value="Verificar Pagamentos")
                            print(f"Linha {row[0].row}: Escreveu Verificar Pagamentos")
                        else:
                            ws.cell(row=row[0].row, column=9, value="Verificar")
                            print(f"Linha {row[0].row}: Escreveu Verificar")
                    else:
                        valor_comparacao = valor_coluna_h_float
                        print(f"Comparando H={valor_comparacao} com {valor}")
                        ws.cell(row=row[0].row, column=10, value=f"{valor_nome}: H={valor_comparacao} vs {valor_nome}={valor}")
                        if valor_comparacao is not None and valor is not None and abs(valor_comparacao - valor) <= tolerancia:
                            ws.cell(row=row[0].row, column=9, value="OK")
                            print(f"Linha {row[0].row}: Escreveu OK (diferença: {abs(valor_comparacao - valor)})")
                        else:
                            ws.cell(row=row[0].row, column=9, value="Verificar")
                            diferenca = abs(valor_comparacao - valor) if valor_comparacao is not None and valor is not None else 'N/A'
                            print(f"Linha {row[0].row}: Escreveu Verificar (diferença: {diferenca})")
                except NameError:
                    print(f"Erro: {valor_nome} não definido. Continuando a execução...")
                    ws.cell(row=row[0].row, column=9, value="Verificar")
                    ws.cell(row=row[0].row, column=10, value=f"Erro: {valor_nome} não definido")
                except KeyError:
                    print(f"Erro: {valor_nome} não encontrado. Continuando a execução...")
                    ws.cell(row=row[0].row, column=9, value="Verificar")
                    ws.cell(row=row[0].row, column=10, value=f"Erro: {valor_nome} não encontrado")
        except TypeError:
            continue

# Salvar as alterações de volta no arquivo Excel
wb.save(excel_path)

time.sleep(5)

# Caminho para o outro script
#caminho_outro_script = f"C:\\patrimonio\\PATRIMONIO.PY"

# subprocess.run([
#     "python", caminho_outro_script,
#     company_code, month_year, company_name
# ])
